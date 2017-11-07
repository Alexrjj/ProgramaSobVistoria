# -*- coding: utf-8 -*-
import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
try:
    from itertools import izip
except ImportError:  # python3.x
    izip = zip

url = 'http://gomnet.ampla.com/'
username = ''
password = ''
wb = openpyxl.load_workbook('sobs.xlsx')
wb1 = openpyxl.load_workbook('sobs.xlsx')

driver = webdriver.Chrome()

if __name__ == '__main__':
    driver.get(url)
    # Faz login no sistema
    uname = driver.find_element_by_name('txtBoxLogin')
    uname.send_keys(username)
    passw = driver.find_element_by_name('txtBoxSenha')
    passw.send_keys(password)
    submit_button = driver.find_element_by_id('ImageButton_Login').click()

    for sheet in wb.worksheets:
        # Busca o menu "Obras" e acessa o submenu "Acompanhamento de Obra"
        menu = driver.find_element_by_class_name('ctl00_Menu_GomNet_3')
        hidden_submenu = driver.find_element_by_xpath('//*[@id="ctl00_Menu_GomNetn6"]/td/table/tbody/tr/td/a')
        webdriver.ActionChains(driver).move_to_element(menu).click(hidden_submenu).perform()

        # Insere o valor na textbox "Número SOB" e realiza a consulta
        sob = driver.find_element_by_name('ctl00$ContentPlaceHolder1$TextBox_NumSOB')
        try:
            sob.send_keys(sheet['A1'].value)
        except TypeError:
            print("Não há sob a ser programada. Fim da execução.")
            break

        driver.find_element_by_id('ctl00_ContentPlaceHolder1_ImageButton_Enviar').click()

        # Clica no botão "Programação da Obra"
        try:
            compel = driver.find_element_by_xpath("//*[contains(text(), 'COMPEL CONSTRUÇÕES MONTAGENS E')]")
            webdriver.ActionChains(driver).click(compel).perform()

            print("Linha encontrada")
            m = 0
            while m <= 8:
                webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
                m += 1
            webdriver.ActionChains(driver).send_keys(Keys.SPACE).perform()
        except NoSuchElementException:
            print("Obra não despachada para a COMPEL. Impossível continuar.")
            break

        # Modifica o atributo CSS da textbox para torná-la editável e insere o valor da variável codTurma
        turma = driver.find_element_by_css_selector('#ctl00_ContentPlaceHolder1_txtBoxTurma')
        driver.execute_script("arguments[0].setAttribute('onkeydown','return true;')", turma)
        turma.send_keys(str(sheet['E1'].value))

        servico = driver.find_element_by_css_selector('#ctl00_ContentPlaceHolder1_txtBoxRespServico')
        driver.execute_script("arguments[0].setAttribute('onkeydown','return true;')", servico)
        servico.send_keys(str(sheet['F1'].value))

        servicoSup = driver.find_element_by_css_selector('#ctl00_ContentPlaceHolder1_txtBoxServicoSuplente')
        driver.execute_script("arguments[0].setAttribute('onkeydown','return true;')", servicoSup)
        servicoSup.send_keys(str(sheet['F1'].value))

        # Procura a textbox "Data Inicio Previsto" e atribui à variável dataInicial
        dataInicial = driver.find_element_by_id('ctl00_ContentPlaceHolder1_Control_DataHora_InicioPrevisto_TextBox6')

        # Move o cursor intermitente para o início da textbox, por já estar pré-formatada
        c = 0
        while c <= 5:
            dataInicial.send_keys(Keys.CONTROL + Keys.LEFT)
            c += 1
        # Busca o valor da variável dataIni, e insere na textbox "Data Inicio Previsto" todos os caracteres,
        for character in str(sheet['B1'].value):
            dataInicial.send_keys(character)

        # Procura a textbox "Data Término Previsto" e atribui à variável dataFinal
        dataFinal = driver.find_element_by_id('ctl00_ContentPlaceHolder1_Control_DataHora_TerminoPrevisto_TextBox6')

        # Move o cursor intermitente para o início da textbox, por já estar pré-formatada
        c = 0
        while c <= 5:
            dataFinal.send_keys(Keys.CONTROL + Keys.LEFT)
            c += 1

        # Busca o valor da variável dataFinal, e insere na textbox "Data Término Previsto" todos os caracteres,
        for character in str(sheet['B2'].value):
            dataFinal.send_keys(character)

        # Adiciona técnicos à tarefa
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_imgBtnGravarTecnicos').click()
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_gridViewTecnicos_ctl01_ChkBoxAll').click()
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_Button1').click()
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_btnVoltar').click()

        # Identifica o menu "Tipo de Programação" e seleciona a opção "Execução de Obra"
        tipoProg = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_TipoProgramacao'))
        tipoProg.select_by_visible_text('EXECUÇÃO DE OBRA')

        # Identifica o menu "Necessita Linha Viva" e seleciona a opção "Não"
        linhaViva = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_linhaViva'))
        linhaViva.select_by_visible_text('NÃO')

        # Identifica o menu "Necessita Desligamento" e seleciona a opção "Não"
        desliga = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_Desliga'))
        desliga.select_by_visible_text('NÃO')

        # Clica no botão "Programar"
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_Button_ProgramarTarefa').click()

        # Preenche o campo "Atividade" com o número da SOB
        atividade = driver.find_element_by_id('ctl00_ContentPlaceHolder1_txtBoxAtividade')
        atividade.send_keys(sheet['A1'].value)

        # Identifica o menu "Horas" e seleciona a opção "01"
        hora = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_Hora'))
        hora.select_by_visible_text('01')

        # Identifica o menu "Minutos" e seleciona a opção "00"
        minuto = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_Minuto'))
        minuto.select_by_visible_text('00')

        # Identifica o menu "Viagem" e seleciona a opção "Não"
        viagem = Select(driver.find_element_by_id('ctl00_ContentPlaceHolder1_DropDownList_Viagem'))
        viagem.select_by_visible_text('NÃO')

        # Clica no botão "Adicionar Programação"
        driver.find_element_by_id('ctl00_ContentPlaceHolder1_btnAdicionarProgramacao').click()
        row = 0
        col = 0
        for (baremo, qtd) in zip(sheet.iter_cols(min_col=3, max_col=3), sheet.iter_cols(min_col=4, max_col=4)):
            for (cell, cell2) in zip(baremo, qtd):
                try: # Procura os baremos na planilha "sobs.xlsx" e marca de acordo
                    driver.find_element_by_xpath("*//tr/td[contains(text(), '" + str(cell.value) + "')]/preceding-sibling::td/input").click()
                    webdriver.ActionChains(driver).send_keys(Keys.TAB).perform()
                    webdriver.ActionChains(driver).send_keys(str(cell2.value)).perform()
                except NoSuchElementException:  # Caso não encontre, abre o arquivo txt e registra o código baremo e sua quantidade
                    log = open("BaremosPendentes.txt", "a")
                    log.write(str(sheet['A1'].value) + " " + str(cell.value) + " " + str(cell2.value) + "\n")
                    log.close()
                continue
        #  Ao fim do loop de inserção de baremos, clica no botão "registrar programação"
        driver.find_element_by_xpath('//*[@id="ctl00_ContentPlaceHolder1_btnEnviarItens"]').click()
        print(str(sheet['A1'].value) + " programada com êxito.")
